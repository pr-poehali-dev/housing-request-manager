"""
Генерация отчётов по заявкам ЖКХ.
POST {action: "excel"} — Excel (.xlsx) base64
POST {action: "pdf"}   — PDF base64
Доступно только диспетчерам.
"""
import json, os, base64, io
from datetime import datetime
import psycopg2

SCHEMA = os.environ.get("MAIN_DB_SCHEMA", "public")

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type, X-Auth-Token",
}

STATUS_LABELS = {
    "new": "Новая", "assigned": "Назначена",
    "in_progress": "В работе", "done": "Выполнена", "waiting": "Ожидает"
}
PRIORITY_LABELS = {
    "urgent": "Срочно", "high": "Высокий", "medium": "Средний", "low": "Низкий"
}

def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"], options=f"-c search_path={SCHEMA}")

def ok(data, status=200):
    return {"statusCode": status, "headers": {**CORS, "Content-Type": "application/json"},
            "body": json.dumps(data, ensure_ascii=False, default=str)}

def err(msg, status=400):
    return {"statusCode": status, "headers": {**CORS, "Content-Type": "application/json"},
            "body": json.dumps({"error": msg}, ensure_ascii=False)}

def get_user(cur, token):
    if not token:
        return None
    cur.execute(
        f"SELECT u.id, u.name, u.role FROM {SCHEMA}.sessions s JOIN {SCHEMA}.users u ON u.id = s.user_id WHERE s.token = %s",
        (token,)
    )
    row = cur.fetchone()
    return {"id": row[0], "name": row[1], "role": row[2]} if row else None

def fetch_requests(cur, date_from=None, date_to=None, status_filter=None):
    sql = f"""
        SELECT r.id, r.category, r.description, r.address, r.status, r.priority,
               r.created_at, r.closed_at, r.close_comment,
               res.name AS resident, res.phone,
               m.name AS master
        FROM {SCHEMA}.requests r
        JOIN {SCHEMA}.users res ON res.id = r.resident_id
        LEFT JOIN {SCHEMA}.users m ON m.id = r.master_id
        WHERE 1=1
    """
    vals = []
    if date_from:
        sql += " AND r.created_at >= %s"; vals.append(date_from)
    if date_to:
        sql += " AND r.created_at <= %s"; vals.append(date_to)
    if status_filter and status_filter != "all":
        sql += " AND r.status = %s"; vals.append(status_filter)
    sql += " ORDER BY r.created_at DESC"
    cur.execute(sql, vals)
    cols = ["id","category","description","address","status","priority",
            "created_at","closed_at","close_comment","resident","phone","master"]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def generate_excel(rows, period_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт по заявкам"

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Отчёт по заявкам ЖКХ — {period_label}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}   Всего заявок: {len(rows)}   Выполнено: {sum(1 for r in rows if r['status']=='done')}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Stats row
    stats = {"Новая": 0, "Назначена": 0, "В работе": 0, "Ожидает": 0, "Выполнена": 0}
    for r in rows:
        lbl = STATUS_LABELS.get(r["status"], r["status"])
        if lbl in stats: stats[lbl] += 1

    ws.merge_cells("A3:L3")
    stat_str = "   |   ".join(f"{k}: {v}" for k, v in stats.items())
    ws["A3"] = stat_str
    ws["A3"].font = Font(size=9, color="444444")
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].fill = PatternFill("solid", fgColor="F0F4FF")

    # Header
    headers = ["№ Заявки","Категория","Адрес","Жилец","Телефон","Мастер","Статус","Приоритет","Создана","Закрыта","Время (ч)","Комментарий диспетчера"]
    header_row = 5
    header_fill = PatternFill("solid", fgColor="1a56db")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    # Status colors
    status_colors = {
        "done": "D1FAE5", "in_progress": "FEF3C7", "assigned": "EDE9FE",
        "waiting": "F1F5F9", "new": "EFF6FF"
    }

    for ri, r in enumerate(rows, header_row + 1):
        created = r["created_at"]
        closed = r["closed_at"]
        if isinstance(created, str):
            try: created = datetime.fromisoformat(created.replace("Z",""))
            except: created = None
        if isinstance(closed, str):
            try: closed = datetime.fromisoformat(closed.replace("Z",""))
            except: closed = None

        hours = ""
        if created and closed:
            delta = closed - created
            hours = round(delta.total_seconds() / 3600, 1)

        status_lbl = STATUS_LABELS.get(r["status"], r["status"])
        priority_lbl = PRIORITY_LABELS.get(r["priority"], r["priority"])
        row_fill = PatternFill("solid", fgColor=status_colors.get(r["status"], "FFFFFF"))

        values = [
            f"№{r['id']}", r["category"], r["address"], r["resident"], r["phone"] or "",
            r["master"] or "—", status_lbl, priority_lbl,
            created.strftime("%d.%m.%Y %H:%M") if created else "",
            closed.strftime("%d.%m.%Y %H:%M") if closed else "—",
            hours, r["close_comment"] or ""
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin
            cell.font = Font(size=9)

    # Column widths
    widths = [10, 18, 22, 18, 14, 16, 12, 10, 16, 16, 9, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = f"A{header_row+1}"

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode()

def generate_pdf(rows, period_label):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import urllib.request

    # Загружаем шрифт с поддержкой кириллицы
    font_url = "https://cdn.jsdelivr.net/npm/@fontsource/pt-sans@5.0.8/files/pt-sans-cyrillic-400-normal.woff2"
    font_path = "/tmp/PTSans.ttf"
    fallback_url = "https://github.com/googlefonts/roboto/raw/main/src/hinted/Roboto-Regular.ttf"

    try:
        urllib.request.urlretrieve("https://fonts.gstatic.com/s/ptsans/v17/jizaRExUiTo99u79D0KEwA.ttf", font_path)
        pdfmetrics.registerFont(TTFont("PTSans", font_path))
        urllib.request.urlretrieve("https://fonts.gstatic.com/s/ptsans/v17/jizfRExUiTo99u79B_mh0O6tKA.ttf", "/tmp/PTSans-Bold.ttf")
        pdfmetrics.registerFont(TTFont("PTSans-Bold", "/tmp/PTSans-Bold.ttf"))
        base_font = "PTSans"
        bold_font = "PTSans-Bold"
    except Exception:
        base_font = "Helvetica"
        bold_font = "Helvetica-Bold"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontName=bold_font, fontSize=14, spaceAfter=4, alignment=1)
    sub_style = ParagraphStyle("sub", fontName=base_font, fontSize=9, textColor=colors.grey, spaceAfter=12, alignment=1)
    cell_style = ParagraphStyle("cell", fontName=base_font, fontSize=8, leading=10)

    done_count = sum(1 for r in rows if r["status"] == "done")
    story = [
        Paragraph(f"Отчёт по заявкам ЖКХ — {period_label}", title_style),
        Paragraph(f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  Всего: {len(rows)}  |  Выполнено: {done_count}", sub_style),
    ]

    headers = ["№","Категория","Адрес","Жилец","Мастер","Статус","Приоритет","Создана","Закрыта","Время","Комментарий"]
    col_widths = [1.2*cm, 3.5*cm, 4.5*cm, 3.5*cm, 3.2*cm, 2.5*cm, 2.2*cm, 3*cm, 3*cm, 1.8*cm, 5*cm]

    table_data = [[Paragraph(h, ParagraphStyle("hdr", fontName=bold_font, fontSize=8, textColor=colors.white)) for h in headers]]

    status_colors_pdf = {
        "done": colors.HexColor("#D1FAE5"),
        "in_progress": colors.HexColor("#FEF3C7"),
        "assigned": colors.HexColor("#EDE9FE"),
        "waiting": colors.HexColor("#F1F5F9"),
        "new": colors.HexColor("#EFF6FF"),
    }
    row_styles = []

    for ri, r in enumerate(rows, 1):
        created = r["created_at"]
        closed = r["closed_at"]
        if isinstance(created, str):
            try: created = datetime.fromisoformat(created.replace("Z",""))
            except: created = None
        if isinstance(closed, str):
            try: closed = datetime.fromisoformat(closed.replace("Z",""))
            except: closed = None

        hours = ""
        if created and closed:
            delta = closed - created
            hours = f"{round(delta.total_seconds()/3600, 1)}ч"

        row_color = status_colors_pdf.get(r["status"], colors.white)
        row_styles.append(("BACKGROUND", (0, ri), (-1, ri), row_color))

        table_data.append([
            Paragraph(f"№{r['id']}", cell_style),
            Paragraph(r["category"] or "", cell_style),
            Paragraph(r["address"] or "", cell_style),
            Paragraph(r["resident"] or "", cell_style),
            Paragraph(r["master"] or "—", cell_style),
            Paragraph(STATUS_LABELS.get(r["status"], r["status"]), cell_style),
            Paragraph(PRIORITY_LABELS.get(r["priority"], r["priority"]), cell_style),
            Paragraph(created.strftime("%d.%m %H:%M") if created else "", cell_style),
            Paragraph(closed.strftime("%d.%m %H:%M") if closed else "—", cell_style),
            Paragraph(hours, cell_style),
            Paragraph(r["close_comment"] or "", cell_style),
        ])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a56db")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUND", (0, 1), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ] + row_styles

    table.setStyle(TableStyle(base_style))
    story.append(table)

    doc.build(story)
    return base64.b64encode(buf.getvalue()).decode()

def handler(event: dict, context) -> dict:
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    headers = event.get("headers") or {}
    token = headers.get("X-Auth-Token") or headers.get("x-auth-token")

    body = {}
    if event.get("body"):
        try:
            body = json.loads(event["body"])
        except Exception:
            return err("Неверный JSON")

    conn = get_conn()
    cur = conn.cursor()
    user = get_user(cur, token)
    if not user:
        return err("Требуется авторизация", 401)
    if user["role"] != "dispatcher":
        return err("Только диспетчер может скачивать отчёты", 403)

    action = body.get("action") or "excel"
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    status_filter = body.get("status_filter") or "all"

    rows = fetch_requests(cur, date_from, date_to, status_filter)

    now = datetime.now().strftime("%d.%m.%Y")
    period_label = f"{date_from or '...'} — {date_to or now}"

    if action == "excel":
        data = generate_excel(rows, period_label)
        fname = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return ok({"data": data, "filename": fname, "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})

    if action == "pdf":
        data = generate_pdf(rows, period_label)
        fname = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return ok({"data": data, "filename": fname, "mime": "application/pdf"})

    return err("Неверное действие", 400)
